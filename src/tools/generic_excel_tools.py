# generic_excel_tools.py
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from langchain_core.tools import tool

from core.tools import instrument_langfuse_tool
from typing import List, Dict, Any, Optional, Union
from datetime import datetime
import shutil

def create_backup(file_path: str) -> str:
    """
    Creates a backup of the Excel file before modification.
    Creates a new backup for every change with a unique timestamp.
    Returns the backup file path.
    """
    # Normalize the file path
    file_path = os.path.abspath(file_path)
    
    try:
        # Create backups directory if it doesn't exist
        backup_dir = os.path.join(os.path.dirname(file_path), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        
        # Generate backup filename with timestamp (including microseconds for uniqueness)
        base_name = os.path.basename(file_path)
        name_without_ext = os.path.splitext(base_name)[0]
        ext = os.path.splitext(base_name)[1]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]  # Include milliseconds
        backup_filename = f"{name_without_ext}_backup_{timestamp}{ext}"
        backup_path = os.path.join(backup_dir, backup_filename)
        
        # Copy the file
        shutil.copy2(file_path, backup_path)
        
        print(f"Backup created: {backup_path}")
        return backup_path
    
    except Exception as e:
        print(f"Warning: Could not create backup: {e}")
        return None

def find_local_status_column(sheet, row_idx: int) -> Optional[int]:
    """
    Finds the Status column for a specific row by looking for the nearest header row above it.
    This is designed for unstructured sheets with multiple tables.
    """
    # Look upward from the current row to find the nearest header row
    for header_row_idx in range(row_idx - 1, 0, -1):
        row_values = [str(cell.value or '').lower() for cell in sheet[header_row_idx]]
        # Check if this looks like a header row (contains "status" and at least one other expected header)
        if "status" in row_values and any(h in row_values for h in ["descriere scurta", "functionalitate", "functionality"]):
            # Find the column index of "status"
            for col_idx, cell in enumerate(sheet[header_row_idx], 1):
                if cell.value and "status" in str(cell.value).lower():
                    return col_idx
            break
    return None

def copy_cell_format_from_above(sheet, row_idx: int, copy_all_columns: bool = True, skip_fill: bool = True):
    """
    Copies formatting (alignment, font, borders) from the row above for each cell in the current row.
    If copy_all_columns is True, copies formatting for all columns, not just those with values.
    If skip_fill is True, does not copy fill/background colors.
    """
    if row_idx <= 1:
        return  # No row above to copy from
    
    max_col = sheet.max_column if copy_all_columns else sheet.max_column
    
    for col_idx in range(1, max_col + 1):
        current_cell = sheet.cell(row=row_idx, column=col_idx)
        above_cell = sheet.cell(row=row_idx - 1, column=col_idx)
        
        # Copy formatting regardless of whether the current cell has a value
        # This ensures borders are copied even to empty cells
        
        # Copy alignment
        if above_cell.alignment:
            current_cell.alignment = Alignment(
                horizontal=above_cell.alignment.horizontal,
                vertical=above_cell.alignment.vertical,
                text_rotation=above_cell.alignment.text_rotation,
                wrap_text=above_cell.alignment.wrap_text,
                shrink_to_fit=above_cell.alignment.shrink_to_fit,
                indent=above_cell.alignment.indent
            )
        
        # Copy font
        if above_cell.font:
            current_cell.font = Font(
                name=above_cell.font.name,
                size=above_cell.font.size,
                bold=above_cell.font.bold,
                italic=above_cell.font.italic,
                vertAlign=above_cell.font.vertAlign,
                underline=above_cell.font.underline,
                strike=above_cell.font.strike,
                color=above_cell.font.color
            )
        
        # Copy borders - this is especially important
        if above_cell.border:
            current_cell.border = Border(
                left=Side(
                    border_style=above_cell.border.left.border_style,
                    color=above_cell.border.left.color
                ) if above_cell.border.left else None,
                right=Side(
                    border_style=above_cell.border.right.border_style,
                    color=above_cell.border.right.color
                ) if above_cell.border.right else None,
                top=Side(
                    border_style=above_cell.border.top.border_style,
                    color=above_cell.border.top.color
                ) if above_cell.border.top else None,
                bottom=Side(
                    border_style=above_cell.border.bottom.border_style,
                    color=above_cell.border.bottom.color
                ) if above_cell.border.bottom else None,
                diagonal=Side(
                    border_style=above_cell.border.diagonal.border_style,
                    color=above_cell.border.diagonal.color
                ) if above_cell.border.diagonal else None,
                diagonal_direction=above_cell.border.diagonal_direction,
                outline=above_cell.border.outline,
                diagonalUp=above_cell.border.diagonalUp,
                diagonalDown=above_cell.border.diagonalDown
            )
        
        # Copy number format
        if above_cell.number_format:
            current_cell.number_format = above_cell.number_format
        
        # SKIP copying fill/background colors when skip_fill is True
        if not skip_fill and above_cell.fill and above_cell.fill.fill_type:
            current_cell.fill = PatternFill(
                fill_type=above_cell.fill.fill_type,
                start_color=above_cell.fill.start_color,
                end_color=above_cell.fill.end_color
            )

def apply_row_style_based_on_status(sheet, row_idx: int, status_col_idx: Optional[int] = None, is_new_row: bool = True, old_status_value: Optional[str] = None):
    """
    Applies appropriate styling to a row based on its status value.
    
    Args:
        sheet: The worksheet object
        row_idx: Row index to apply styling to
        status_col_idx: Column index of the status column (optional)
        is_new_row: True if this is a newly inserted row, False if updating existing row
        old_status_value: The previous status value (only used when is_new_row=False)
    """
    # For new rows, copy formatting from the row above (excluding fill colors)
    if is_new_row:
        copy_cell_format_from_above(sheet, row_idx, copy_all_columns=True, skip_fill=True)
    
    if status_col_idx is None:
        # Try to find the local Status column for this row
        status_col_idx = find_local_status_column(sheet, row_idx)
    
    if status_col_idx is None:
        print(f"Warning: Could not find Status column for row {row_idx}")
        return
    
    # Get the current status value
    status_cell = sheet.cell(row=row_idx, column=status_col_idx)
    status_value = str(status_cell.value or "").lower().strip()
    
    # For updates, check if status has actually changed
    if not is_new_row and old_status_value is not None:
        old_status_normalized = str(old_status_value or "").lower().strip()
        if status_value == old_status_normalized:
            # Status hasn't changed, don't modify any styling
            return
    
    # Determine styling based on status
    fill = None
    font_color = None
    font_bold = None
    
    if "done" in status_value:
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        font_color = "006100"
        font_bold = True
    elif "in testing" in status_value or "testing" in status_value:
        fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        font_color = "9C5700"
        font_bold = False
    elif "partial" in status_value:
        fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_color = "9C0006"
        font_bold = False
    else:
        # For empty or unrecognized status values, remove styling
        # Clear fill for all cells in the row
        for cell in sheet[row_idx]:
            if cell.value is not None:
                # Remove fill (set to no fill)
                cell.fill = PatternFill(fill_type=None)
                
                # Reset font to default while preserving basic properties
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=False,  # Reset to not bold
                        italic=cell.font.italic,
                        vertAlign=cell.font.vertAlign,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color="000000"  # Reset to black
                    )
        return  # Exit early since we've removed styling
    
    # Apply the styling to all cells in the row with content
    if fill and font_color is not None:
        for cell in sheet[row_idx]:
            if cell.value is not None:
                # Apply fill
                cell.fill = fill
                
                # Update font while preserving other properties
                if cell.font:
                    cell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=font_bold if font_bold is not None else cell.font.bold,
                        italic=cell.font.italic,
                        vertAlign=cell.font.vertAlign,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=font_color
                    )

@tool
def update_cell_in_unstructured_sheet(file_path: str, search_value: str, column_header: Union[str, int], new_value: Any) -> str:
    """
    Performs a robust update on a single cell in a sheet with multiple, unstructured tables (like the 'BID' file).
    
    Args:
        file_path: Path to the Excel file
        search_value: Text to search for in the row
        column_header: Either a column header name (string) or a column position (integer, 1-based)
        new_value: The new value to set in the cell
    
    When column_header is an integer, it directly uses that column position.
    When column_header is a string, it searches for that header in the table.
    """
    print(f"--- TOOL CALLED: update_cell_in_unstructured_sheet (robust method) for search '{search_value}' ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Step 1: Find the target row
        target_row_idx = -1
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), 1):
            row_text = ' '.join([str(c.value or '') for c in row])
            if search_value.lower() in row_text.lower():
                target_row_idx = row_idx
                break
        
        if target_row_idx == -1:
            return f"Error: Could not find a row containing '{search_value}'."

        # Handle column_header as either string (header name) or integer (column position)
        target_col_idx = -1
        status_col_idx = None
        
        # If column_header is an integer, use it directly as column index
        if isinstance(column_header, int):
            target_col_idx = column_header
            # Still try to find status column for styling
            status_col_idx = find_local_status_column(sheet, target_row_idx)
        else:
            # Original behavior: search for header
            # Step 2: Find the nearest header row for the table this row belongs to
            header_row_idx = -1
            header_map = {}
            for idx in range(target_row_idx - 1, 0, -1):
                current_row = sheet[idx]
                # A row is considered a header if it contains at least two of the expected labels
                potential_headers = [str(cell.value or '').lower() for cell in current_row]
                if any(h in potential_headers for h in ["status", "descriere scurta", "functionalitate"]):
                    header_row_idx = idx
                    # Step 3: Build the precise header-to-column map from this specific row
                    for cell in current_row:
                        if cell.value:
                            header_map[str(cell.value)] = cell.column
                    break
            
            if not header_map and isinstance(column_header, str):
                return f"Error: Could not find a valid header row above the row containing '{search_value}'."

            # Step 4: Find the correct column index from our precise map
            # Allow for partial matches in keys, e.g., 'Status' should match 'Status'
            for header_text, col_idx in header_map.items():
                if column_header.lower() in header_text.lower():
                    target_col_idx = col_idx
                if "status" in header_text.lower():
                    status_col_idx = col_idx
            
            if target_col_idx == -1:
                found_headers = ", ".join(header_map.keys())
                return f"Error: Column '{column_header}' not found in the table's local headers. Found headers: [{found_headers}]."

        # Validate column index
        if target_col_idx < 1 or target_col_idx > sheet.max_column:
            return f"Error: Column index {target_col_idx} is out of range. Sheet has {sheet.max_column} columns."

        # Get the old status value if we're updating the status column
        old_status_value = None
        if status_col_idx and ((isinstance(column_header, str) and "status" in column_header.lower()) or 
                               (isinstance(column_header, int) and column_header == status_col_idx)):
            old_status_value = sheet.cell(row=target_row_idx, column=status_col_idx).value
        elif status_col_idx:
            # Get the current status value even if we're not updating status column
            old_status_value = sheet.cell(row=target_row_idx, column=status_col_idx).value

        # Step 5: Perform the surgical update and provide a detailed success message
        cell_to_update = sheet.cell(row=target_row_idx, column=target_col_idx)
        original_value = cell_to_update.value
        cell_to_update.value = new_value
        
        # Apply styling based on the status column, but only if status has changed
        apply_row_style_based_on_status(sheet, target_row_idx, status_col_idx, is_new_row=False, old_status_value=old_status_value)
        
        workbook.save(file_path)

        styling_note = ""
        if isinstance(column_header, str) and "status" in column_header.lower() and str(original_value or "").lower().strip() != str(new_value or "").lower().strip():
            styling_note = " Row styling has been updated based on the new status."
        
        column_desc = f"column {target_col_idx}" if isinstance(column_header, int) else f"local header '{column_header}' in column {get_column_letter(target_col_idx)}"
        
        return (f"Update successful. In row {target_row_idx}, found the {column_desc}. "
                f"Changed cell {get_column_letter(target_col_idx)}{target_row_idx} from '{original_value}' to '{new_value}'.{styling_note} "
                f"Backup created: {backup_path}")
    except Exception as e:
        return f"An unexpected error occurred during robust update: {e}"

@tool
def insert_row_after_anchor(file_path: str, anchor_text: str, row_data: List[str], position: str = "below") -> str:
    """
    Inserts a new row above or below a row containing the anchor text.
    This actually creates a new row and shifts existing content down, properly handling merged cells.
    Automatically applies styling based on Status column values and copies formatting.
    Creates a backup before making changes.
    
    Args:
        file_path: Path to the Excel file
        anchor_text: Text to search for in the Excel file
        row_data: List of values to insert in the new row
        position: "above" or "below" - where to insert relative to anchor (default: "below")
    """
    print(f"--- TOOL CALLED: insert_row_after_anchor, anchoring to '{anchor_text}', position: {position} ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    if position not in ["above", "below"]:
        return f"Error: position must be 'above' or 'below', got '{position}'"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        # Find the anchor row
        anchor_row_idx = -1
        anchor_col_idx = 1  # Default to column A
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value and anchor_text.lower() in str(cell.value).lower():
                    anchor_row_idx = cell.row
                    anchor_col_idx = cell.column
                    break
            if anchor_row_idx != -1:
                break
        
        if anchor_row_idx == -1:
            return f"Error: Could not find anchor text '{anchor_text}' in the file."
        
        # Determine where to insert
        if position == "below":
            insert_row_idx = anchor_row_idx + 1
        else:  # above
            insert_row_idx = anchor_row_idx
        
        # Handle merged cells before insertion
        merges_to_restore = []
        for merged_range in list(sheet.merged_cells.ranges):
            # If the merged range is at or below the insertion point, it needs to be adjusted
            if merged_range.min_row >= insert_row_idx:
                merges_to_restore.append({
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                    "value": sheet.cell(merged_range.min_row, merged_range.min_col).value
                })
                sheet.unmerge_cells(str(merged_range))
        
        # Insert a new row at the specified position
        sheet.insert_rows(insert_row_idx)
        
        # Add the data to the new row
        for i, cell_value in enumerate(row_data):
            sheet.cell(row=insert_row_idx, column=anchor_col_idx + i, value=cell_value)
        
        # Apply styling based on status (this is a new row)
        apply_row_style_based_on_status(sheet, insert_row_idx, is_new_row=True)
        
        # Restore merged cells with adjusted row positions
        for merge_info in merges_to_restore:
            new_min_row = merge_info["min_row"] + 1
            new_max_row = merge_info["max_row"] + 1
            new_range_str = (
                f"{get_column_letter(merge_info['min_col'])}{new_min_row}:"
                f"{get_column_letter(merge_info['max_col'])}{new_max_row}"
            )
            sheet.merge_cells(new_range_str)
            # Restore the value in the top-left cell of the merged range
            sheet.cell(row=new_min_row, column=merge_info["min_col"]).value = merge_info["value"]
        
        workbook.save(file_path)
        
        return f"Successfully inserted a new row {position} '{anchor_text}' at row {insert_row_idx}, starting at column {get_column_letter(anchor_col_idx)}. Merged cells have been preserved and formatting applied. Backup created: {backup_path}"
    
    except Exception as e:
        return f"An unexpected error occurred while inserting row: {e}"

@tool
def insert_row_at_position(file_path: str, sheet_name: str, row_number: int, row_data: List[str]) -> str:
    """
    Inserts a new row at a specific row number in a named sheet.
    Existing rows at and below this position are shifted down, properly handling merged cells.
    Automatically applies styling based on Status column values and copies formatting.
    Creates a backup before making changes.
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet to insert into
        row_number: Row number where to insert (1-based)
        row_data: List of values to insert in the new row
    """
    print(f"--- TOOL CALLED: insert_row_at_position at row {row_number} in sheet '{sheet_name}' ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    if row_number < 1:
        return f"Error: Row number must be 1 or greater, got {row_number}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        
        if sheet_name not in workbook.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available sheets are: {', '.join(workbook.sheetnames)}."
        
        sheet = workbook[sheet_name]
        
        # Handle merged cells before insertion
        merges_to_restore = []
        for merged_range in list(sheet.merged_cells.ranges):
            # If the merged range is at or below the insertion point, it needs to be adjusted
            if merged_range.min_row >= row_number:
                merges_to_restore.append({
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                    "value": sheet.cell(merged_range.min_row, merged_range.min_col).value
                })
                sheet.unmerge_cells(str(merged_range))
        
        # Insert a new row at the specified position
        sheet.insert_rows(row_number)
        
        # Add the data to the new row
        for col_idx, cell_value in enumerate(row_data, 1):
            sheet.cell(row=row_number, column=col_idx, value=cell_value)
        
        # Apply styling based on status (this is a new row)
        apply_row_style_based_on_status(sheet, row_number, is_new_row=True)
        
        # Restore merged cells with adjusted row positions
        for merge_info in merges_to_restore:
            new_min_row = merge_info["min_row"] + 1
            new_max_row = merge_info["max_row"] + 1
            new_range_str = (
                f"{get_column_letter(merge_info['min_col'])}{new_min_row}:"
                f"{get_column_letter(merge_info['max_col'])}{new_max_row}"
            )
            sheet.merge_cells(new_range_str)
            # Restore the value in the top-left cell of the merged range
            sheet.cell(row=new_min_row, column=merge_info["min_col"]).value = merge_info["value"]
        
        workbook.save(file_path)
        
        return f"Successfully inserted a new row at position {row_number} in sheet '{sheet_name}'. Merged cells have been preserved and formatting applied. Backup created: {backup_path}"
    
    except Exception as e:
        return f"An unexpected error occurred while inserting row: {e}"

@tool
def apply_styling_to_all_rows(file_path: str) -> str:
    """
    Applies appropriate styling to all rows in the Excel file based on their Status values.
    This is a refresh operation that applies consistent styling across the entire file.
    Creates a backup before making changes.
    """
    print(f"--- TOOL CALLED: apply_styling_to_all_rows ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        sheets_processed = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows_styled = 0
            
            # Apply styling to each row
            for row_idx in range(1, sheet.max_row + 1):
                # Check if this row has content
                row_has_content = any(cell.value for cell in sheet[row_idx])
                if row_has_content:
                    # Check if it's a header row (skip styling for headers)
                    row_values = [str(cell.value or '').lower() for cell in sheet[row_idx]]
                    is_header = "status" in row_values and any(h in row_values for h in ["descriere scurta", "functionalitate", "functionality"])
                    
                    if not is_header:
                        # For refresh operations, we treat all rows as existing rows
                        # but we don't have old status values, so styling will always be applied
                        apply_row_style_based_on_status(sheet, row_idx, is_new_row=False, old_status_value=None)
                        rows_styled += 1
            
            sheets_processed.append(f"{sheet_name} ({rows_styled} rows styled)")
        
        workbook.save(file_path)
        return f"Successfully applied styling and formatting to all rows based on status. Sheets processed: {', '.join(sheets_processed)}. Backup created: {backup_path}"
    
    except Exception as e:
        return f"An unexpected error occurred while applying styling: {e}"

@tool
def find_and_add_row_in_excel(file_path: str, anchor_text: str, row_data: List[str], insert_mode: bool = True) -> str:
    """
    Adds a new row under a section identified by `anchor_text`.
    Automatically applies styling based on Status column values and copies formatting.
    Creates a backup before making changes.
    
    Args:
        file_path: Path to the Excel file
        anchor_text: Text to search for as the anchor point
        row_data: List of values to add
        insert_mode: If True, inserts a new row. If False, finds empty row (legacy behavior)
    """
    print(f"--- TOOL CALLED: find_and_add_row_in_excel, anchoring to '{anchor_text}', insert_mode: {insert_mode} ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        
        anchor_row_idx = -1
        anchor_col_idx = -1
        
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                if cell.value and anchor_text.lower() in str(cell.value).lower():
                    anchor_row_idx = cell.row
                    anchor_col_idx = cell.column
                    break
            if anchor_row_idx != -1:
                break
        
        if anchor_row_idx == -1:
            return f"Error: Could not find anchor text '{anchor_text}'."
        
        if insert_mode:
            # Insert a new row below the anchor
            insert_row_idx = anchor_row_idx + 1
            
            # Handle merged cells before insertion
            merges_to_restore = []
            for merged_range in list(sheet.merged_cells.ranges):
                # If the merged range is at or below the insertion point, it needs to be adjusted
                if merged_range.min_row >= insert_row_idx:
                    merges_to_restore.append({
                        "min_row": merged_range.min_row,
                        "max_row": merged_range.max_row,
                        "min_col": merged_range.min_col,
                        "max_col": merged_range.max_col,
                        "value": sheet.cell(merged_range.min_row, merged_range.min_col).value
                    })
                    sheet.unmerge_cells(str(merged_range))
            
            sheet.insert_rows(insert_row_idx)
            
            # Add the data to the new row
            for i, cell_value in enumerate(row_data):
                sheet.cell(row=insert_row_idx, column=anchor_col_idx + i, value=cell_value)
            
            # Apply styling based on status (this is a new row)
            apply_row_style_based_on_status(sheet, insert_row_idx, is_new_row=True)
            
            # Restore merged cells with adjusted row positions
            for merge_info in merges_to_restore:
                new_min_row = merge_info["min_row"] + 1
                new_max_row = merge_info["max_row"] + 1
                new_range_str = (
                    f"{get_column_letter(merge_info['min_col'])}{new_min_row}:"
                    f"{get_column_letter(merge_info['max_col'])}{new_max_row}"
                )
                sheet.merge_cells(new_range_str)
                # Restore the value in the top-left cell of the merged range
                sheet.cell(row=new_min_row, column=merge_info["min_col"]).value = merge_info["value"]
            
            workbook.save(file_path)
            return f"Successfully inserted a new row below '{anchor_text}' at row {insert_row_idx}, starting at column {anchor_col_idx}. Merged cells have been preserved and formatting applied. Backup created: {backup_path}"
        else:
            # Legacy behavior: find empty row
            insert_row_idx = anchor_row_idx + 1
            while sheet.cell(row=insert_row_idx, column=anchor_col_idx).value is not None:
                insert_row_idx += 1
                if insert_row_idx > sheet.max_row + 5:
                    return f"Error: Could not find an empty row below the anchor."
            
            for i, cell_value in enumerate(row_data):
                sheet.cell(row=insert_row_idx, column=anchor_col_idx + i, value=cell_value)
            
            # Apply styling based on status (this is a new row)
            apply_row_style_based_on_status(sheet, insert_row_idx, is_new_row=True)
            
            workbook.save(file_path)
            return f"Successfully found section '{anchor_text}' and added the new row at row {insert_row_idx}, starting at column {anchor_col_idx}. Formatting and styling applied. Backup created: {backup_path}"
    
    except Exception as e:
        return f"An unexpected error occurred: {e}"

@tool
def update_row(file_path: str, sheet_name: str, search_value: str, new_data: Dict[str, Any]) -> str:
    """
    Use this tool ONLY for structured sheets where headers are in the FIRST row.
    Only applies styling if the Status column value has changed.
    Creates a backup before making changes.
    """
    print(f"--- TOOL CALLED: update_row on sheet '{sheet_name}' ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available sheets are: {', '.join(workbook.sheetnames)}."
        sheet = workbook[sheet_name]
        headers = {cell.value: cell.column for cell in sheet[1]}
        target_row_idx = -1
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            row_text = ' '.join([str(c.value or '') for c in row])
            if search_value.lower() in row_text.lower():
                target_row_idx = row_idx
                break
        if target_row_idx == -1:
            return f"Error: Could not find a row containing the value '{search_value}'."
        
        # Check if we're updating the status column and get old status value
        old_status_value = None
        status_col_idx = None
        if "Status" in headers:
            status_col_idx = headers["Status"]
            old_status_value = sheet.cell(row=target_row_idx, column=status_col_idx).value
        
        # Perform the updates
        for col_name, new_value in new_data.items():
            if col_name not in headers:
                return f"Error: Column '{col_name}' not found in the sheet headers."
            col_idx = headers[col_name]
            sheet.cell(row=target_row_idx, column=col_idx).value = new_value
        
        # Apply styling based on status changes
        apply_row_style_based_on_status(sheet, target_row_idx, status_col_idx, is_new_row=False, old_status_value=old_status_value)
        
        workbook.save(file_path)
        
        styling_note = ""
        if "Status" in new_data and str(old_status_value or "").lower().strip() != str(new_data["Status"] or "").lower().strip():
            styling_note = " Styling applied based on status change."
            
        return f"Successfully updated the row identified by '{search_value}' in sheet '{sheet_name}'.{styling_note} Backup created: {backup_path}"
    except Exception as e:
        return f"An unexpected error occurred during update: {e}"

@tool
def delete_row(file_path: str, sheet_name: str, search_value: str) -> str:
    """
    Performs a "smart" deletion. It truly removes the row and shifts cells up, while intelligently preserving the sheet's structure, including merged cells.
    Creates a backup before making changes.
    """
    print(f"--- TOOL CALLED: delete_row (smart shift method) on sheet '{sheet_name}' ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]
        target_row_idx = -1
        for row_idx, row in enumerate(sheet.iter_rows(), 1):
            row_text = ' '.join([str(c.value or '') for c in row])
            if search_value.lower() in row_text.lower():
                target_row_idx = row_idx
                break
        if target_row_idx == -1:
            return f"Error: Could not find a row containing '{search_value}'."
        merges_to_restore = []
        for merged_range in list(sheet.merged_cells.ranges):
            if merged_range.min_row > target_row_idx:
                merges_to_restore.append({
                    "min_row": merged_range.min_row, "min_col": merged_range.min_col,
                    "max_row": merged_range.max_row, "max_col": merged_range.max_col,
                    "value": merged_range.start_cell.value
                })
                sheet.unmerge_cells(str(merged_range))
        sheet.delete_rows(target_row_idx)
        for merge_info in merges_to_restore:
            new_min_row = merge_info["min_row"] - 1
            new_max_row = merge_info["max_row"] - 1
            new_range_str = (
                f"{get_column_letter(merge_info['min_col'])}{new_min_row}:"
                f"{get_column_letter(merge_info['max_col'])}{new_max_row}"
            )
            sheet.merge_cells(new_range_str)
            top_left_cell = sheet.cell(row=new_min_row, column=merge_info["min_col"])
            top_left_cell.value = merge_info["value"]
        workbook.save(file_path)
        return f"Successfully performed a smart deletion of the row identified by '{search_value}' from sheet '{sheet.title}'. Backup created: {backup_path}"
    except Exception as e:
        return f"An unexpected error occurred during smart deletion: {e}"

@tool
def list_sheets(file_path: str) -> str:
    """Lists all sheet names in the Excel file."""
    print(f"--- TOOL CALLED: list_sheets for file '{file_path}' ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    try:
        workbook = load_workbook(file_path, read_only=True)
        return f"The available sheets are: {', '.join(workbook.sheetnames)}"
    except Exception as e:
        return f"An error occurred while listing sheets: {e}"

@tool
def inspect_excel_file(file_path: str) -> str:
    """Inspects an Excel file to understand its structure, such as sheet names and column headers."""
    print(f"--- TOOL CALLED: inspect_excel_file for file '{file_path}' ---")
    if not os.path.exists(file_path):
        return f"Error: File not found at path: {file_path}"
    try:
        xls = pd.ExcelFile(file_path)
        if not xls.sheet_names: return "The Excel file is empty and contains no sheets."
        summary_lines = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name).fillna('') 
            summary_lines.append(f"--- Sheet: '{sheet_name}' ---")
            summary_lines.append(f"Columns: {df.columns.to_list()}")
            summary_lines.append("First 5 rows of data:")
            summary_lines.append(df.head().to_markdown(index=False))
        return "\n\n".join(summary_lines)
    except Exception as e:
        return f"Error reading Excel file: {e}."

@tool
def answer_question_from_excel(file_path: str, question: str) -> str:
    """Answers a specific question about the content of an Excel file by reading the entire file."""
    print(f"--- TOOL CALLED: answer_question_from_excel for question '{question}' ---")
    if not os.path.exists(file_path): return f"Error: File not found at path: {file_path}"
    try:
        xls = pd.ExcelFile(file_path)
        full_context = []
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet_name).fillna('')
            full_context.append(f"--- Content from Sheet: '{sheet_name}' ---\n{df.to_markdown(index=False)}")
        context_str = "\n".join(full_context)
        return f"CONTEXT:\n{context_str}\n\nBased on the context, answer the question: {question}"
    except Exception as e:
        return f"Error reading Excel file: {e}"

@tool
def add_row_to_named_sheet(file_path: str, sheet_name: str, row_data: List[str]) -> str:
    """
    Adds a new row to the end of a SPECIFICALLY NAMED sheet. Use for structured tables.
    Automatically applies styling based on Status column values and copies formatting.
    Creates a backup before making changes.
    """
    print(f"--- TOOL CALLED: add_row_to_named_sheet to sheet '{sheet_name}' ---")
    if not os.path.exists(file_path): return f"Error: File not found at path: {file_path}"
    
    # Create backup before modification
    backup_path = create_backup(file_path)
    
    try:
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found. Available sheets are: {', '.join(workbook.sheetnames)}."
        sheet = workbook[sheet_name]
        sheet.append(row_data)
        
        # Apply styling to the newly added row
        new_row_idx = sheet.max_row
        apply_row_style_based_on_status(sheet, new_row_idx, is_new_row=True)
        
        workbook.save(file_path)
        return f"Successfully added the new row to sheet '{sheet_name}' with appropriate styling and formatting. Backup created: {backup_path}"
    except Exception as e:
        return f"An unexpected error occurred: {e}"


# Instrument Langfuse telemetry for all generic Excel tools
update_cell_in_unstructured_sheet = instrument_langfuse_tool(
    update_cell_in_unstructured_sheet, name="update_cell_in_unstructured_sheet"
)
insert_row_after_anchor = instrument_langfuse_tool(
    insert_row_after_anchor, name="insert_row_after_anchor"
)
insert_row_at_position = instrument_langfuse_tool(
    insert_row_at_position, name="insert_row_at_position"
)
apply_styling_to_all_rows = instrument_langfuse_tool(
    apply_styling_to_all_rows, name="apply_styling_to_all_rows"
)
find_and_add_row_in_excel = instrument_langfuse_tool(
    find_and_add_row_in_excel, name="find_and_add_row_in_excel"
)
update_row = instrument_langfuse_tool(update_row, name="update_row")
delete_row = instrument_langfuse_tool(delete_row, name="delete_row")
list_sheets = instrument_langfuse_tool(list_sheets, name="list_sheets")
inspect_excel_file = instrument_langfuse_tool(
    inspect_excel_file, name="inspect_excel_file"
)
answer_question_from_excel = instrument_langfuse_tool(
    answer_question_from_excel, name="answer_question_from_excel"
)
add_row_to_named_sheet = instrument_langfuse_tool(
    add_row_to_named_sheet, name="add_row_to_named_sheet"
)
