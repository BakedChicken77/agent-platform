import os
from langchain_core.tools import BaseTool, tool
from core import settings
import openpyxl


def write_to_xlsx(data: list) -> str:
    """
    Tool: Writes a list of rows (each row is a list or tuple) into an Excel file.
    Validates that each row matches the expected schema length.
    """
    global expected_excel_header
    sheet_name = 'Sheet'
    output_file = 'temp.xlsx'

    # Check for empty or malformed data
    if not data or not isinstance(data, list) or not all(isinstance(row, list) for row in data):
        return "Error: Invalid data format. Must be a list of lists."

    # Validate that each row matches the header length
    for i, row in enumerate(data):
        if len(row) != len(expected_excel_header):
            return (
                f"Error: Row {i + 1} does not match expected column count.\n"
                f"Expected {len(expected_excel_header)} columns: {expected_excel_header}\n"
                f"Received {len(row)} columns: {row}"
            )

    # Load or create workbook
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Append all data rows
    for row_data in data:
        ws.append(row_data)

    wb.save(output_file)
    return f"Data written to {output_file} in sheet {sheet_name}."

write_to_excel: BaseTool = tool(write_to_xlsx)
write_to_excel.name = "write_to_excel_xlsx_file"  
